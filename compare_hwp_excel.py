# -*- coding: utf-8 -*-
"""
compare_hwp_excel.py
=====================
한글(HWP) 문서를 "기준(source of truth)"으로 삼아 엑셀 파일과 비교하고,
엑셀을 한글 내용에 맞게 동일화(sync)하는 스크립트.

비교 대상 4개 항목 (요청사항 그대로):
    1) 한글 "제품 및 물질"      <-> 엑셀 "제품 및 물질"
    2) 한글 "규격번호"          <-> 엑셀 "규격코드"
    3) 한글 "규격명"            <-> 엑셀 "규격명"
    4) 한글 "구성요소,특성(시험범위)" <-> 엑셀 "구성요소,특성(시험범위)(한글)"

매칭(짝짓기) 방법
------------------
행을 매칭할 고유 키가 마땅치 않은 경우가 많으므로, 기본적으로
"규격번호(=규격코드)"를 키로 사용합니다. 규격번호가 같은데 나머지
값이 다르면 "내용 불일치"로, 규격번호 자체가 한쪽에만 있으면
"누락/추가"로 분류합니다.

[2026-08 수정 1] 소재지(사업장) 매칭 버그 수정
--------------------------------------------
기존 location_same()이 정규화된 소재지 코드끼리 "포함관계(in)"까지
동일한 것으로 판정하고 있었습니다. 그런데 실제 데이터에는 "소재지"
(번호 없음)와 "소재지-1", "소재지-2" ... 처럼 서로 다른 독립된 값이
공존하기 때문에, 문자열 "소재지"가 "소재지-2"의 부분 문자열이 되어
전혀 다른 사업장(예: 진주 vs 안산)이 같은 곳으로 오판정되는 문제가
있었습니다. 이번 수정으로 소재지 코드는 완전일치(==)만 허용합니다.

[2026-08 수정 2] 동일 엑셀 행 중복 매칭 버그 수정  ← 이번에 추가로 발견/수정
--------------------------------------------------
compare() 의 "1차(정확 키) 매칭" 경로에서 이미 다른 HWP 레코드가
사용한 엑셀 행(used_excel_indices)을 걸러내지 않고 있었습니다.
그 결과 (규격번호+소재지) 조합이 HWP 쪽에 두 번 이상 등장하면
(페이지 분할 표 등으로 동일 항목이 중복 추출되는 경우), 서로 다른
HWP 레코드 여러 개가 엑셀의 같은 행 하나에 중복으로 매칭되고,
엑셀에 실제로 존재하는 나머지 행은 "엑셀에만 있음(한글 없음)"으로
잘못 표시되는 문제가 있었습니다. 이번 수정으로 1차 매칭에서도
이미 사용된 엑셀 행은 후보에서 제외합니다.

[2026-08 수정 3] HWP 전용/엑셀 전용 리스트에 소재지 표기 추가
--------------------------------------------------------------
"한글에는 있는데 엑셀에 없는 규격번호" / "엑셀에는 있는데 한글에 없는
규격번호" 목록에 소재지 정보가 빠져 있어, 동일한 규격번호가 사업장별로
서로 다르게 누락된 경우(예: 진주 사업장만 누락, 안산 사업장은 정상)를
구분할 수 없었습니다. 이번 수정으로 HTML/마크다운 리포트의 두 목록
모두에 소재지를 함께 표기합니다.

사용법
------
    python compare_hwp_excel.py \
        --hwp "2_최종_인정신청분야_및_범위-_주_엠아이티지.hwp" \
        --xls "_주_엠아이티지_인정분야.xls" \
        --out-dir ./output

결과물 (out-dir 안에 생성):
    - diff_report.md         : 사람이 읽기 좋은 비교 리포트 (Markdown)
    - diff_report.csv        : 같은 내용을 엑셀에서 열어볼 수 있는 CSV
    - diff_report.html       : 색상으로 차이를 강조한 HTML 리포트
    - synced.xlsx             : 한글 내용대로 값이 보정된 엑셀 (변경 셀 노란색 강조)

Claude API를 활용한 유사도 판정 (선택 사항)
--------------------------------------------
공백/기호 차이만 있는데 "다르다"고 잘못 판정되는 것을 줄이기 위해,
--use-claude 옵션을 주면 애매한 케이스(문자열이 다른데 얼마나 다른지
애매한 경우)에 대해 Anthropic Claude API를 호출해서
"의미상 동일한지" 여부를 한 번 더 확인합니다.
    - 환경변수 ANTHROPIC_API_KEY 필요
    - pip install anthropic
"""

from __future__ import annotations

import argparse
import csv
import difflib
import html
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from hwp_parser import parse_hwp_main_table, parse_hwp_all_tables


# --------------------------------------------------------------------------
# 1. 엑셀 컬럼 -> 표준 이름 매핑
#    (엑셀 헤더 문구가 프로젝트마다 조금씩 다를 수 있어 키워드 방식으로 탐색)
# --------------------------------------------------------------------------
EXCEL_HEADER_KEYWORDS = {
    "product": ["제품", "물질"],
    "spec_no": ["규격코드"],
    "spec_name": ["규격명"],
    "component": ["구성요소", "특성"],
}

# 소재지(사업장)는 파일마다 "사업장", "사업장구분", "소재지" 등으로 헤더명이
# 다를 수 있고, 없는 파일도 있을 수 있어서 필수 컬럼과 분리해서 관리한다.
# (이 목록에 있는 키워드 중 하나라도 포함되면 매칭)
LOCATION_HEADER_KEYWORDS = ["사업장", "소재지"]

# 비교 결과 각 필드에 대응하는, synced.xlsx 에 실제로 값을 써넣을 엑셀 컬럼 후보
# (엑셀 헤더 자동탐색 실패 시 이 이름으로 폴백)
FALLBACK_EXCEL_COLS = {
    "product": "제품 및 물질",
    "spec_no": "규격코드",
    "spec_name": "규격명",
    "component": "구성요소,특성(시험범위)(한글)",
}
FALLBACK_LOCATION_COL = "사업장구분"


def find_excel_columns(columns: List[str]) -> Dict[str, Optional[str]]:
    col_map: Dict[str, Optional[str]] = {k: None for k in EXCEL_HEADER_KEYWORDS}
    for col in columns:
        col_str = str(col)
        norm = col_str.replace(" ", "")
        for std_name, keywords in EXCEL_HEADER_KEYWORDS.items():
            if col_map[std_name] is not None:
                continue
            # "구성요소,특성(시험범위)(영문)" 같은 영문 컬럼은 제외
            if "영문" in norm:
                continue
            if all(kw in norm for kw in keywords):
                col_map[std_name] = col_str
    # 폴백 처리
    for k, v in col_map.items():
        if v is None and FALLBACK_EXCEL_COLS[k] in columns:
            col_map[k] = FALLBACK_EXCEL_COLS[k]

    # 소재지(사업장) 컬럼은 선택 사항 -> 못 찾아도 에러 내지 않고 None으로 둠
    location_col: Optional[str] = None
    for col in columns:
        col_str = str(col)
        norm = col_str.replace(" ", "")
        if "영문" in norm:
            continue
        if any(kw in norm for kw in LOCATION_HEADER_KEYWORDS):
            location_col = col_str
            break
    if location_col is None and FALLBACK_LOCATION_COL in columns:
        location_col = FALLBACK_LOCATION_COL
    col_map["location"] = location_col
    return col_map


# --------------------------------------------------------------------------
# 2. 텍스트 정규화 (비교용) - 공백/괄호/기호 차이로 인한 오탐 줄이기
# --------------------------------------------------------------------------
def normalize(text: object) -> str:
    if text is None:
        return ""
    s = str(text)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    # 전각/반각, 유사 기호 통일 (예: 물결표, 콜론 앞뒤 공백 등)
    s = s.replace("：", ":").replace("～", "~")
    # 괄호/슬래시 바로 안쪽의 불필요한 공백 제거 (예: "( IEC ...)" vs "(IEC ...)",
    # "㎍ /L" vs "㎍/L") -> 내용은 같은데 공백 차이만으로 오탐(❌) 나는 것을 방지
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    s = re.sub(r"\s+/", "/", s)
    s = re.sub(r"/\s+", "/", s)
    # 콜론 앞뒤 공백 표기 통일 (예: "Test Load :" vs "Test Load:")
    s = re.sub(r"\s*:\s*", ": ", s)
    # "제 1부" vs "제1부", "제 2-2 부" vs "제2-2부" 처럼 "제"와 숫자(하이픈 포함)
    # 사이, 숫자와 "부/절/항/편" 사이의 공백 표기 차이를 통일
    s = re.sub(r"제\s+([\d\-\.]+)", r"제\1", s)          # "제 " -> "제"
    s = re.sub(r"([\d\-\.]+)\s+(부|절|항|편)", r"\1\2", s)  # "N.N 절" -> "N.N절"
    # 하이픈 계열(엔대시/엠대시) 표기 통일
    s = s.replace("–", "-").replace("—", "-")
    s = s.strip()
    return s


def normalize_spec_no(text: object) -> str:
    """규격번호/규격코드는 공백을 아예 없애고 대문자로 통일해서 매칭 키로 사용."""
    s = normalize(text)
    s = s.upper().replace(" ", "")
    return s


def normalize_location(text: object) -> str:
    """
    소재지/사업장 비교용 정규화.

    한글(HWP) 소제목은 "소재지-1, 구로"처럼 [코드 + 지역명]이 함께 있지만,
    엑셀 "사업장" 컬럼은 실무상 "소재지-1"처럼 [코드]만 들어있는 경우가
    대부분이라, 뒤에 붙는 지역명은 버리고 앞의 코드 부분만 뽑아 비교 키로
    쓴다. (코드 패턴을 못 찾으면 전체 텍스트를 공백 제거해서 그대로 사용)

    주의: "소재지"(번호 없음)와 "소재지-1", "소재지-2" ... 는 실제로는
    서로 다른, 독립된 사업장 코드다. \\d* 로 인해 "소재지"만 있어도
    매치되지만, 반환되는 코드 자체는 "소재지"와 "소재지-N"으로 서로
    다르게 나오므로 이후 비교는 반드시 완전일치(==)로 판정해야 한다
    (location_same 참고).
    """
    s = normalize(text)
    m = re.match(r"^(소재지\s*-?\s*\d*|부속시설\s*-?\s*\d+)", s)
    if m:
        code = re.sub(r"\s+", "", m.group(1))
        # "소재지1" 처럼 하이픈이 빠진 표기도 "소재지-1"로 통일
        code = re.sub(r"^(소재지|부속시설)(\d)", r"\1-\2", code)
        return code
    return s.replace(" ", "")


def location_same(hwp_loc: str, excel_loc: str) -> bool:
    """
    소재지 코드가 완전히 일치하는지 판정한다.

    [수정] 기존에는 normalize_location() 결과끼리 포함관계(in)까지
    동일한 것으로 허용했다 (a in b or b in a). 그런데 실제 데이터에는
    "소재지"(번호 없음)와 "소재지-1" / "소재지-2" ... 가 서로 다른
    독립된 사업장으로 공존하며, "소재지"라는 문자열이 "소재지-1",
    "소재지-2" 등의 부분 문자열이 되어 버려 전혀 다른 사업장(예: 진주
    vs 안산)이 같은 곳으로 오판정되는 문제가 있었다.

    normalize_location()이 이미 코드를 깔끔하게 추출해주므로, 여기서는
    완전일치만 허용한다.
    """
    a, b = normalize_location(hwp_loc), normalize_location(excel_loc)
    if not a or not b:
        return False
    return a == b


# --------------------------------------------------------------------------
# 3. 비교 결과 자료구조
# --------------------------------------------------------------------------
@dataclass
class FieldDiff:
    field: str
    hwp_value: str
    excel_value: str
    same: bool


@dataclass
class RowResult:
    spec_no_key: str
    status: str  # "matched" | "hwp_only" | "excel_only"
    hwp_record: Optional[Dict[str, str]] = None
    excel_row_index: Optional[int] = None  # 엑셀 데이터프레임 상의 행 index
    field_diffs: List[FieldDiff] = field(default_factory=list)
    excel_location: str = ""  # excel_only 상태일 때 표시용 소재지 원문
    excel_spec_no_raw: str = ""  # excel_only 상태일 때 표시용 규격코드 원문(정규화 전)

    @property
    def has_mismatch(self) -> bool:
        return any(not d.same for d in self.field_diffs)


# --------------------------------------------------------------------------
# 4. 메인 비교 로직
# --------------------------------------------------------------------------
def load_excel(xls_path: str) -> Tuple[pd.DataFrame, Dict[str, Optional[str]]]:
    engine = "xlrd" if xls_path.lower().endswith(".xls") else None
    df = pd.read_excel(xls_path, engine=engine, dtype=str)
    df = df.fillna("")
    col_map = find_excel_columns(list(df.columns))
    # "location"(소재지/사업장)은 선택 항목이라 필수 컬럼 누락 검사에서 제외
    missing = [k for k, v in col_map.items() if v is None and k != "location"]
    if missing:
        raise RuntimeError(
            f"엑셀에서 다음 컬럼을 찾지 못했습니다: {missing}\n"
            f"실제 컬럼 목록: {list(df.columns)}"
        )
    if col_map.get("location") is None:
        print(
            "[안내] 엑셀에서 소재지/사업장 컬럼을 찾지 못해 소재지 비교는 건너뜁니다.",
            file=sys.stderr,
        )
    return df, col_map


def compare(hwp_records: List[Dict[str, str]], df: pd.DataFrame, col_map: Dict[str, str]) -> List[RowResult]:
    location_col = col_map.get("location")
    # 소재지 매칭을 켤지 여부: 엑셀에 소재지 컬럼이 있고, 한글 쪽 레코드에도
    # location 정보가 하나라도 채워져 있으면(=parse_hwp_all_tables로 파싱했으면) 사용.
    use_location = bool(location_col) and any(rec.get("location") for rec in hwp_records)

    def make_key(spec_no_norm: str, location_raw: str) -> str:
        if use_location:
            return f"{spec_no_norm}||{normalize_location(location_raw)}"
        return spec_no_norm

    # 엑셀: 매칭 키 -> 행 index 리스트 (중복 있을 수 있음)
    excel_key_to_indices: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        spec_key = normalize_spec_no(row[col_map["spec_no"]])
        loc_val = str(row[location_col]) if location_col else ""
        key = make_key(spec_key, loc_val)
        excel_key_to_indices.setdefault(key, []).append(idx)

    results: List[RowResult] = []
    used_excel_indices = set()

    for rec in hwp_records:
        spec_key = normalize_spec_no(rec.get("spec_no", ""))
        if not spec_key:
            continue
        hwp_loc = rec.get("location", "")
        key = make_key(spec_key, hwp_loc)

        # [수정 2] 1차(정확 키) 매칭에서도 이미 다른 HWP 레코드가 사용한
        # 엑셀 행은 후보에서 제외한다. 이게 빠져 있으면 (규격번호+소재지)가
        # HWP 쪽에 중복으로 등장할 때 같은 엑셀 행에 여러 번 매칭되고,
        # 엑셀에 실제로 존재하는 나머지 행이 "엑셀에만 있음"으로 잘못
        # 표시되는 문제가 생긴다.
        candidates = [
            idx for idx in excel_key_to_indices.get(key, [])
            if idx not in used_excel_indices
        ]
        if not candidates and use_location:
            # 소재지까지 붙여서 못 찾았으면, 소재지 표기 차이(코드 정규식
            # 추출이 안 되는 비정형 텍스트 등)까지 감안해서 같은 규격번호를
            # 가진 후보들 중 소재지가 "동일 코드"로 판정되는 것을 한 번 더
            # 찾아본다. (location_same은 완전일치만 허용하므로, 여기서
            # "소재지" vs "소재지-2"처럼 서로 다른 사업장이 잘못 엮이지
            # 않는다)
            spec_only_candidates = [
                idx for idx, row in df.iterrows()
                if normalize_spec_no(row[col_map["spec_no"]]) == spec_key
                and idx not in used_excel_indices
            ]
            candidates = [
                idx for idx in spec_only_candidates
                if location_same(hwp_loc, str(df.loc[idx, location_col]))
            ]

        if not candidates:
            results.append(RowResult(spec_no_key=key, status="hwp_only", hwp_record=rec))
            continue

        excel_idx = candidates[0]  # 동일 키가 여러 개면 첫 번째만 사용
        used_excel_indices.add(excel_idx)
        row = df.loc[excel_idx]

        diffs = [
            FieldDiff(
                field="제품 및 물질",
                hwp_value=rec.get("product", ""),
                excel_value=str(row[col_map["product"]]),
                same=normalize(rec.get("product", "")) == normalize(row[col_map["product"]]),
            ),
            FieldDiff(
                field="규격번호/규격코드",
                hwp_value=rec.get("spec_no", ""),
                excel_value=str(row[col_map["spec_no"]]),
                same=normalize_spec_no(rec.get("spec_no", "")) == normalize_spec_no(row[col_map["spec_no"]]),
            ),
            FieldDiff(
                field="규격명",
                hwp_value=rec.get("spec_name", ""),
                excel_value=str(row[col_map["spec_name"]]),
                same=normalize(rec.get("spec_name", "")) == normalize(row[col_map["spec_name"]]),
            ),
            FieldDiff(
                field="구성요소,특성(시험범위)",
                hwp_value=rec.get("component", ""),
                excel_value=str(row[col_map["component"]]),
                same=normalize(rec.get("component", "")) == normalize(row[col_map["component"]]),
            ),
        ]

        if location_col:
            excel_loc = str(row[location_col])
            diffs.append(
                FieldDiff(
                    field="소재지/사업장",
                    hwp_value=hwp_loc,
                    excel_value=excel_loc,
                    # hwp 쪽에 소재지 정보가 아예 없는 표(주사업장 등)는
                    # 비교 대상에서 제외하고 "일치"로 처리 (오탐 방지)
                    same=(not hwp_loc) or location_same(hwp_loc, excel_loc),
                )
            )

        results.append(
            RowResult(
                spec_no_key=key,
                status="matched",
                hwp_record=rec,
                excel_row_index=excel_idx,
                field_diffs=diffs,
            )
        )

    # 엑셀에만 있고 한글엔 없는 행들
    for idx, row in df.iterrows():
        if idx in used_excel_indices:
            continue
        spec_key = normalize_spec_no(row[col_map["spec_no"]])
        loc_val = str(row[location_col]) if location_col else ""
        key = make_key(spec_key, loc_val)
        results.append(RowResult(
            spec_no_key=key,
            status="excel_only",
            excel_row_index=idx,
            excel_location=loc_val,
            excel_spec_no_raw=str(row[col_map["spec_no"]]),
        ))

    return results


# --------------------------------------------------------------------------
# 5. (선택) Claude API로 애매한 텍스트 차이 재판정
# --------------------------------------------------------------------------
def refine_with_claude(results: List[RowResult]) -> None:
    """
    normalize() 만으로는 "다르다"고 나오지만, 실제로는 의미가 같은
    (띄어쓰기/기호 차이 정도) 애매한 케이스를 Claude에게 다시 물어봐서
    same 여부를 보정한다. field_diffs 를 in-place로 수정.
    """
    try:
        import anthropic
    except ImportError:
        print("[경고] anthropic 패키지가 없어 --use-claude 를 건너뜁니다. "
              "`pip install anthropic` 후 다시 시도하세요.", file=sys.stderr)
        return

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("[경고] ANTHROPIC_API_KEY 환경변수가 없어 --use-claude 를 건너뜁니다.",
              file=sys.stderr)
        return

    client = anthropic.Anthropic(api_key=api_key)

    for r in results:
        if r.status != "matched":
            continue
        for d in r.field_diffs:
            if d.same:
                continue
            # 완전히 다른 문자열(길이 차이가 너무 큰 경우)은 굳이 물어보지 않음 -> 비용 절감
            if not d.hwp_value or not d.excel_value:
                continue
            len_ratio = len(d.hwp_value) / max(1, len(d.excel_value))
            if len_ratio < 0.5 or len_ratio > 2.0:
                continue

            prompt = (
                "다음 두 텍스트는 시험 규격/성분 관련 문서에서 나온 값입니다. "
                "표기나 단위, 띄어쓰기 차이가 아니라 '실질적인 값(숫자, 단위, 조성 등)'이 "
                "다른지 판단해주세요. 실질적으로 동일하면 SAME, 다르면 DIFFERENT 라고만 "
                "한 단어로 답하세요.\n\n"
                f"텍스트 A (한글 문서 기준):\n{d.hwp_value}\n\n"
                f"텍스트 B (엑셀):\n{d.excel_value}"
            )
            try:
                msg = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=10,
                    messages=[{"role": "user", "content": prompt}],
                )
                answer = "".join(
                    block.text for block in msg.content if block.type == "text"
                ).strip().upper()
                if answer.startswith("SAME"):
                    d.same = True
            except Exception as e:  # 네트워크 오류 등은 조용히 무시하고 원래 판정 유지
                print(f"[경고] Claude 호출 실패, 원래 판정 유지: {e}", file=sys.stderr)


# --------------------------------------------------------------------------
# 6. 리포트 출력
# --------------------------------------------------------------------------
def _safe_path_if_locked(path: str) -> str:
    """
    대상 파일이 다른 프로그램(엑셀 등)에서 열려 있어 쓰기 권한이 없는 경우,
    타임스탬프를 붙인 다른 파일명을 대신 반환한다. (PermissionError 방지)
    """
    if not os.path.exists(path):
        return path
    try:
        # 실제로 쓰기 가능한지 열어서 확인 (내용은 바꾸지 않음)
        with open(path, "a"):
            pass
        return path
    except PermissionError:
        import datetime
        base, ext = os.path.splitext(path)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = f"{base}_{ts}{ext}"
        print(
            f"[경고] '{path}' 파일이 다른 프로그램(엑셀 등)에서 열려 있어 "
            f"'{alt_path}' 이름으로 대신 저장합니다. (원래 파일을 닫고 다시 실행하면 "
            f"동일한 파일명으로 저장할 수 있습니다.)",
            file=sys.stderr,
        )
        return alt_path


def _esc(text: str) -> str:
    """HTML escape 후 줄바꿈은 <br>로 변환 (표 안에서 줄바꿈이 보이도록)."""
    return html.escape(text).replace("\n", "<br>")


def highlight_diff_html(a: str, b: str) -> Tuple[str, str]:
    """
    두 문자열(a=한글 기준, b=엑셀)을 글자 단위로 비교해서, 서로 다른
    부분만 색으로 강조한 HTML 조각을 반환한다.
    - a(한글)쪽: 엑셀에는 없고 한글에만 있는 부분 -> 빨간 취소선
    - b(엑셀)쪽: 한글에는 없고 엑셀에만 있는 부분 -> 초록 배경
    같은 부분은 그대로 표시.
    """
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    out_a: List[str] = []
    out_b: List[str] = []
    DEL = 'style="background-color:#ffd6d6;color:#b30000;text-decoration:line-through;"'
    INS = 'style="background-color:#d6ffd6;color:#006600;"'
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        seg_a, seg_b = _esc(a[i1:i2]), _esc(b[j1:j2])
        if tag == "equal":
            out_a.append(seg_a)
            out_b.append(seg_b)
        elif tag == "delete":
            out_a.append(f"<span {DEL}>{seg_a}</span>")
        elif tag == "insert":
            out_b.append(f"<span {INS}>{seg_b}</span>")
        elif tag == "replace":
            out_a.append(f"<span {DEL}>{seg_a}</span>")
            out_b.append(f"<span {INS}>{seg_b}</span>")
    return "".join(out_a), "".join(out_b)


def write_html_report(results: List[RowResult], out_dir: str) -> str:
    """색상으로 차이를 강조한 HTML 리포트 생성 (브라우저에서 열어서 확인)."""
    html_path = _safe_path_if_locked(os.path.join(out_dir, "diff_report.html"))

    matched = [r for r in results if r.status == "matched"]
    mismatched = [r for r in matched if r.has_mismatch]
    hwp_only = [r for r in results if r.status == "hwp_only"]
    excel_only = [r for r in results if r.status == "excel_only"]

    parts: List[str] = []
    parts.append("<!DOCTYPE html><html lang='ko'><head><meta charset='utf-8'>")
    parts.append("<title>한글(HWP) vs 엑셀 비교 리포트</title>")
    parts.append(
        "<style>"
        "body{font-family:'Malgun Gothic',sans-serif;margin:24px;line-height:1.5;}"
        "table{border-collapse:collapse;width:100%;margin-bottom:24px;table-layout:fixed;}"
        "th,td{border:1px solid #ccc;padding:8px;vertical-align:top;word-break:break-word;}"
        "th{background:#f0f0f0;text-align:left;}"
        "td.field{white-space:nowrap;width:160px;font-weight:bold;background:#fafafa;}"
        "td.mark{width:60px;text-align:center;}"
        "h2{border-bottom:2px solid #333;padding-bottom:4px;margin-top:40px;}"
        "h3{background:#eef;padding:6px 10px;border-left:4px solid #557;}"
        "ul{line-height:1.8;}"
        ".loc{color:#557;font-weight:normal;}"
        "</style></head><body>"
    )
    parts.append("<h1>한글(HWP) vs 엑셀 비교 리포트</h1>")
    parts.append("<ul>")
    parts.append(f"<li>매칭된 규격번호 수: {len(matched)}</li>")
    parts.append(f"<li>내용이 다른 규격번호 수: {len(mismatched)}</li>")
    parts.append(f"<li>한글에만 있는 규격번호 수(엑셀 누락): {len(hwp_only)}</li>")
    parts.append(f"<li>엑셀에만 있는 규격번호 수(한글에 없음): {len(excel_only)}</li>")
    parts.append("</ul>")

    if mismatched:
        parts.append("<h2>내용 불일치 상세</h2>")
        for r in mismatched:
            spec_no = html.escape(r.hwp_record.get("spec_no", ""))
            loc = r.hwp_record.get("location", "")
            title = spec_no + (f" ({html.escape(loc)})" if loc else "")
            parts.append(f"<h3>규격번호: {title}</h3>")
            parts.append(
                "<table><tr><th>항목</th><th>한글(기준)</th><th>엑셀</th><th>일치여부</th></tr>"
            )
            for d in r.field_diffs:
                mark = "✅" if d.same else "❌"
                if d.same:
                    a_html, b_html = _esc(d.hwp_value), _esc(d.excel_value)
                else:
                    # 실제 판정 로직과 동일하게 정규화(공백/줄바꿈 정리)한 값끼리
                    # 비교해서 강조 표시 -> 서식 차이 노이즈 없이 진짜 차이만 색으로 표시
                    norm_field = "spec_no" if d.field == "규격번호/규격코드" else None
                    if norm_field == "spec_no":
                        norm_a, norm_b = normalize_spec_no(d.hwp_value), normalize_spec_no(d.excel_value)
                    else:
                        norm_a, norm_b = normalize(d.hwp_value), normalize(d.excel_value)
                    a_html, b_html = highlight_diff_html(norm_a, norm_b)
                field = html.escape(d.field)
                parts.append(
                    f"<tr><td class='field'>{field}</td>"
                    f"<td>{a_html}</td><td>{b_html}</td>"
                    f"<td class='mark'>{mark}</td></tr>"
                )
            parts.append("</table>")

    if hwp_only:
        parts.append("<h2>한글에는 있는데 엑셀에 없는 규격번호</h2><ul>")
        for r in hwp_only:
            spec_no = html.escape(r.hwp_record.get("spec_no", ""))
            spec_name = html.escape(r.hwp_record.get("spec_name", ""))
            loc = r.hwp_record.get("location", "")
            loc_html = f" <span class='loc'>[{html.escape(loc)}]</span>" if loc else ""
            parts.append(f"<li>{spec_no} : {spec_name}{loc_html}</li>")
        parts.append("</ul>")

    if excel_only:
        parts.append(
            "<h2>엑셀에는 있는데 한글에 없는 규격번호 "
            "(한글 문서에 해당 대/중분류가 없을 수도 있음)</h2><ul>"
        )
        for r in excel_only:
            spec_no_display = r.excel_spec_no_raw or r.spec_no_key
            loc = r.excel_location
            loc_html = f" <span class='loc'>[{html.escape(loc)}]</span>" if loc else ""
            parts.append(f"<li>{html.escape(spec_no_display)}{loc_html}</li>")
        parts.append("</ul>")

    parts.append("</body></html>")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write("".join(parts))

    return html_path


def write_reports(results: List[RowResult], out_dir: str) -> Tuple[str, str, str]:
    os.makedirs(out_dir, exist_ok=True)
    md_path = _safe_path_if_locked(os.path.join(out_dir, "diff_report.md"))
    csv_path = _safe_path_if_locked(os.path.join(out_dir, "diff_report.csv"))

    matched = [r for r in results if r.status == "matched"]
    mismatched = [r for r in matched if r.has_mismatch]
    hwp_only = [r for r in results if r.status == "hwp_only"]
    excel_only = [r for r in results if r.status == "excel_only"]

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# 한글(HWP) vs 엑셀 비교 리포트\n\n")
        f.write(f"- 매칭된 규격번호 수: {len(matched)}\n")
        f.write(f"- 내용이 다른 규격번호 수: {len(mismatched)}\n")
        f.write(f"- 한글에만 있는 규격번호 수(엑셀 누락): {len(hwp_only)}\n")
        f.write(f"- 엑셀에만 있는 규격번호 수(한글에 없음): {len(excel_only)}\n\n")

        if mismatched:
            f.write("## 내용 불일치 상세\n\n")
            for r in mismatched:
                loc = r.hwp_record.get("location", "")
                title = r.hwp_record.get('spec_no', '') + (f" ({loc})" if loc else "")
                f.write(f"### 규격번호: {title}\n\n")
                f.write("| 항목 | 한글(기준) | 엑셀 | 일치여부 |\n")
                f.write("|---|---|---|---|\n")
                for d in r.field_diffs:
                    mark = "✅" if d.same else "❌"
                    f.write(f"| {d.field} | {d.hwp_value} | {d.excel_value} | {mark} |\n")
                f.write("\n")

        if hwp_only:
            f.write("## 한글에는 있는데 엑셀에 없는 규격번호\n\n")
            for r in hwp_only:
                loc = r.hwp_record.get('location', '')
                loc_str = f" [{loc}]" if loc else ""
                f.write(f"- {r.hwp_record.get('spec_no', '')} : {r.hwp_record.get('spec_name', '')}{loc_str}\n")
            f.write("\n")

        if excel_only:
            f.write("## 엑셀에는 있는데 한글에 없는 규격번호 (한글 문서에 해당 대/중분류가 없을 수도 있음)\n\n")
            for r in excel_only:
                spec_no_display = r.excel_spec_no_raw or r.spec_no_key
                loc_str = f" [{r.excel_location}]" if r.excel_location else ""
                f.write(f"- {spec_no_display}{loc_str}\n")
            f.write("\n")

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["규격번호", "상태", "항목", "한글(기준)", "엑셀", "일치여부"])
        for r in results:
            if r.status == "matched":
                for d in r.field_diffs:
                    writer.writerow([
                        r.hwp_record.get("spec_no", ""),
                        "일치" if not r.has_mismatch else "불일치",
                        d.field, d.hwp_value, d.excel_value,
                        "일치" if d.same else "불일치",
                    ])
            elif r.status == "hwp_only":
                writer.writerow([r.hwp_record.get("spec_no", ""), "엑셀누락", "", "", "", ""])
            else:
                writer.writerow([r.spec_no_key, "한글없음", "", "", "", ""])

    html_path = write_html_report(results, out_dir)

    return md_path, csv_path, html_path


# --------------------------------------------------------------------------
# 7. 엑셀 동일화(sync) - 한글 값으로 덮어쓰기 + 변경 셀 노란색 강조
# --------------------------------------------------------------------------
def write_synced_excel(xls_path: str, results: List[RowResult], col_map: Dict[str, str], out_dir: str) -> str:
    # openpyxl은 legacy .xls를 못 읽으므로, pandas로 읽은 뒤 xlsx로 새로 저장
    df = pd.read_excel(xls_path, engine=("xlrd" if xls_path.lower().endswith(".xls") else None), dtype=str)
    df = df.fillna("")

    changed_cells = []  # (row_idx(0-base, 데이터 시작 기준), col_name)

    for r in results:
        if r.status != "matched" or not r.has_mismatch:
            continue
        for d in r.field_diffs:
            std_name = {
                "제품 및 물질": "product",
                "규격번호/규격코드": "spec_no",
                "규격명": "spec_name",
                "구성요소,특성(시험범위)": "component",
                "소재지/사업장": "location",
            }[d.field]
            if d.same:
                continue
            excel_col = col_map.get(std_name)
            if not excel_col:
                continue  # 소재지 컬럼이 엑셀에 없는 경우 등은 건너뜀
            df.at[r.excel_row_index, excel_col] = d.hwp_value
            changed_cells.append((r.excel_row_index, excel_col))

    out_path = _safe_path_if_locked(os.path.join(out_dir, "synced.xlsx"))
    df.to_excel(out_path, index=False)

    # 변경된 셀 노란색 강조 + 폰트 통일
    wb = load_workbook(out_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx, col_name in changed_cells:
        col_idx = header.index(col_name) + 1  # 1-base
        excel_row = row_idx + 2  # 1행은 헤더, pandas index는 0부터, 데이터는 2행부터
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.fill = fill

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=cell.font.size or 10)

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------
# 8. 파일 선택 (경로를 안 주면 탐색기 창을 띄워서 고르게 함)
# --------------------------------------------------------------------------
def pick_file_with_dialog(title: str, filetypes: List[Tuple[str, str]]) -> str:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        raise RuntimeError(
            "tkinter를 사용할 수 없습니다. --hwp/--xls 옵션으로 경로를 직접 지정해주세요."
        )
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    if not path:
        raise SystemExit("파일을 선택하지 않아 종료합니다.")
    return path


def resolve_hwp_xls_from_dropped(paths: List[str]) -> Tuple[str, str]:
    """드래그&드롭으로 받은 파일 경로 목록에서 hwp/xls를 자동으로 구분."""
    hwp_path = next((p for p in paths if p.lower().endswith(".hwp")), None)
    xls_path = next((p for p in paths if p.lower().endswith((".xls", ".xlsx"))), None)
    if not hwp_path or not xls_path:
        raise SystemExit(
            "드래그한 파일 중 .hwp 하나와 .xls/.xlsx 하나가 반드시 있어야 합니다.\n"
            f"받은 파일: {paths}"
        )
    return hwp_path, xls_path


# --------------------------------------------------------------------------
# 9. CLI
# --------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="HWP(기준) vs Excel 비교/동일화 도구")
    parser.add_argument("--hwp", help="기준이 되는 .hwp 파일 경로")
    parser.add_argument("--xls", help="비교할 .xls/.xlsx 파일 경로")
    parser.add_argument("--out-dir", default="./output", help="결과물 저장 폴더")
    parser.add_argument("--use-claude", action="store_true",
                         help="애매한 불일치를 Claude API로 재판정 (ANTHROPIC_API_KEY 필요)")
    # 드래그&드롭용: hwp/xls 파일을 스크립트(또는 .bat) 위에 끌어다 놓으면
    # 여기로 파일 경로가 그대로 인자로 들어온다.
    parser.add_argument("dropped", nargs="*", help=argparse.SUPPRESS)
    args = parser.parse_args()

    if args.hwp and args.xls:
        hwp_path, xls_path = args.hwp, args.xls
    elif args.dropped:
        hwp_path, xls_path = resolve_hwp_xls_from_dropped(args.dropped)
    else:
        # 아무 경로도 안 줬으면 탐색기 창을 띄워서 고르게 함
        print("경로가 지정되지 않아 파일 선택 창을 엽니다...")
        hwp_path = pick_file_with_dialog(
            "기준이 되는 .hwp 파일 선택", [("HWP 파일", "*.hwp"), ("모든 파일", "*.*")]
        )
        xls_path = pick_file_with_dialog(
            "비교할 .xls/.xlsx 파일 선택", [("Excel 파일", "*.xls;*.xlsx"), ("모든 파일", "*.*")]
        )

    print(f"[1/5] HWP 파싱 중: {hwp_path}")
    try:
        # 부속시설별로 표가 여러 개 있으면 전부 합치고, 표 위 "(...)" 제목을
        # 소재지로 함께 붙여서 반환한다. (소재지 비교/복합키 매칭에 사용)
        hwp_records = parse_hwp_all_tables(hwp_path)
    except RuntimeError as e:
        print(f"      -> 표 자동 인식 실패({e}), 가장 큰 표 하나만 사용합니다.")
        hwp_records = parse_hwp_main_table(hwp_path)
    n_with_loc = sum(1 for r in hwp_records if r.get("location"))
    print(f"      -> {len(hwp_records)}개 행 추출 (소재지 정보 있는 행: {n_with_loc}개)")

    print(f"[2/5] 엑셀 로딩 중: {xls_path}")
    df, col_map = load_excel(xls_path)
    print(f"      -> 컬럼 매핑: {col_map}")

    print("[3/5] 비교 중...")
    results = compare(hwp_records, df, col_map)

    if args.use_claude:
        print("[3.5/5] Claude API로 애매한 케이스 재판정 중...")
        refine_with_claude(results)

    print("[4/5] 리포트 작성 중...")
    md_path, csv_path, html_path = write_reports(results, args.out_dir)
    print(f"      -> {md_path}")
    print(f"      -> {csv_path}")
    print(f"      -> {html_path}  (색상으로 차이 강조, 브라우저로 열어보세요)")

    print("[5/5] 엑셀 동일화(synced.xlsx) 작성 중...")
    synced_path = write_synced_excel(xls_path, results, col_map, args.out_dir)
    print(f"      -> {synced_path}")

    print("\n완료.")
    # 드래그&드롭(.bat)으로 실행했을 때 창이 바로 닫히는 것을 방지
    try:
        input("\n결과 확인 후 아무 키나 눌러 창을 닫으세요...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()